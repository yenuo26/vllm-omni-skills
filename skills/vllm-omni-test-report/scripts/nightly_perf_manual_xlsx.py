#!/usr/bin/env python3
"""Load ``nightly_perf_manual.xlsx`` (beside ``nightly_jobs``) for nightly HTML/Markdown reports."""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

PERF_MANUAL_FILENAME = "nightly_perf_manual.xlsx"
# Place next to current workbook to show ↑/↓ % vs previous run (same sheet name, same row/column).
PERF_MANUAL_PREV_FILENAME = "nightly_perf_manual.prev.xlsx"

# Worksheet names excluded from nightly 「性能测试结果」(raw dumps — too large / not for summary tables).
PERF_MANUAL_SHEETS_SKIP: frozenset[str] = frozenset({"omni_raw", "diffusion_raw"})

DEFAULT_MAX_ROWS_PER_SHEET = 500
DEFAULT_MAX_COLS = 72
DEFAULT_MAX_SHEETS = 24


def resolve_perf_manual_path(log_dir: Path) -> Path:
    """
    ``log_dir`` is typically ``.../logs/nightly_jobs``; the workbook is ``.../logs/nightly_perf_manual.xlsx``.
    """
    return (log_dir.resolve().parent) / PERF_MANUAL_FILENAME


def resolve_perf_manual_prev_path(log_dir: Path) -> Path:
    """Sibling of ``nightly_perf_manual.xlsx`` used for ↑/↓ percentage deltas."""
    return (log_dir.resolve().parent) / PERF_MANUAL_PREV_FILENAME


def _parse_loose_number(s: str) -> float | None:
    """Parse a leading or first numeric token; allow trailing unit suffix (e.g. ``ms``, ``MB``)."""
    if not s or not str(s).strip():
        return None
    t = str(s).strip()
    m = re.search(r"[-+]?(?:\d+\.?\d*|\.\d+)(?:[eE][-+]?\d+)?", t)
    if not m:
        return None
    before = t[: m.start()].strip()
    if before:
        return None
    after = t[m.end() :].strip()
    if after.startswith("%"):
        after = after[1:].strip()
    if after and not re.fullmatch(r"[A-Za-zµμΩΩ°%]+$", after):
        return None
    try:
        return float(m.group(0))
    except ValueError:
        return None


def perf_cell_delta_suffix(cur_cell: str, prev_cell: str) -> str:
    """
    Return ``\"\"``, ``\"↑n.n%\"``, or ``\"↓n.n%\"`` for table suffix.

    If display strings are identical, returns ``\"\"``. If both parse as numbers and differ,
    percentage is ``(cur-prev)/|prev|*100`` (one decimal). Otherwise ``\"\"``.
    """
    a = (cur_cell or "").strip()
    b = (prev_cell or "").strip()
    if a == b:
        return ""
    x = _parse_loose_number(a)
    y = _parse_loose_number(b)
    if x is None or y is None:
        return ""
    if y == 0:
        return ""
    pct = (x - y) / abs(y) * 100.0
    if abs(pct) < 1e-6:
        return ""
    if pct > 0:
        return f"↑{pct:.1f}%"
    return f"↓{abs(pct):.1f}%"


def _delta_grid(cur_rows: list[list[str]], prev_rows: list[list[str]]) -> list[list[str]]:
    out: list[list[str]] = []
    for i, crow in enumerate(cur_rows):
        prow = prev_rows[i] if i < len(prev_rows) else []
        row_out: list[str] = []
        for j, c in enumerate(crow):
            pc = prow[j] if j < len(prow) else ""
            row_out.append(perf_cell_delta_suffix(c, pc))
        out.append(row_out)
    return out


def annotate_perf_deltas(current: dict[str, Any], previous: dict[str, Any]) -> None:
    """Mutate ``current`` sheets: add ``delta_rows`` (same shape as ``rows``) when a prev sheet matches."""
    if current.get("status") != "ok" or previous.get("status") != "ok":
        return
    prev_map = {sh["title"]: sh for sh in previous.get("sheets", [])}
    for sh in current.get("sheets", []):
        psh = prev_map.get(sh["title"])
        if not psh:
            sh["delta_rows"] = [[""] * len(r) for r in sh["rows"]]
            continue
        sh["delta_rows"] = _delta_grid(sh["rows"], psh["rows"])


def load_perf_manual_with_compare(log_dir: Path) -> dict[str, Any]:
    """
    Load current ``nightly_perf_manual.xlsx`` and, if ``nightly_perf_manual.prev.xlsx`` exists,
    attach ``delta_rows`` per sheet plus ``compare_path``.
    """
    data = load_perf_manual_sheets(resolve_perf_manual_path(log_dir))
    prev_path = resolve_perf_manual_prev_path(log_dir)
    if data.get("status") == "ok" and prev_path.is_file():
        prev_data = load_perf_manual_sheets(prev_path)
        if prev_data.get("status") == "ok":
            annotate_perf_deltas(data, prev_data)
            data["compare_path"] = str(prev_path.resolve())
            note = (
                f"与上一版对比：`{prev_path.name}`（同工作表名、同行同列；仅当单元格文本不同且两侧均为可解析数字时显示 ↑/↓ 百分比）。"
            )
            existing = (data.get("message") or "").strip()
            data["message"] = f"{existing} {note}".strip() if existing else note
    return data


def _cell_str(val: Any) -> str:
    if val is None:
        return ""
    if isinstance(val, str):
        return val.strip()
    return str(val)


def _normalize_grid(
    rows: list[list[str]],
    max_cols: int,
) -> tuple[list[str], list[list[str]]]:
    if not rows:
        return [], []
    mc = min(max((len(r) for r in rows), default=0), max_cols)
    headers = (rows[0] + [""] * mc)[:mc]
    for i, h in enumerate(headers):
        t = h.strip()
        headers[i] = t if t else f"col{i + 1}"
    body: list[list[str]] = []
    for r in rows[1:]:
        padded = (r + [""] * mc)[:mc]
        body.append([(c[:2000] if len(c) > 2000 else c) for c in padded])
    return headers, body


def load_perf_manual_sheets(
    path: Path,
    *,
    max_rows_per_sheet: int = DEFAULT_MAX_ROWS_PER_SHEET,
    max_cols: int = DEFAULT_MAX_COLS,
    max_sheets: int = DEFAULT_MAX_SHEETS,
) -> dict[str, Any]:
    """
    Returns a dict: ``path``, ``status`` (``ok`` | ``missing`` | ``no_openpyxl`` | ``error``),
    ``message``, ``sheets`` (list of ``title``, ``headers``, ``rows``, ``truncated_rows``).
    """
    path = path.resolve()
    out: dict[str, Any] = {
        "path": str(path),
        "status": "ok",
        "message": "",
        "sheets": [],
    }
    if not path.is_file():
        out["status"] = "missing"
        return out
    try:
        from openpyxl import load_workbook  # type: ignore[import-untyped]
    except ImportError:
        out["status"] = "no_openpyxl"
        out["message"] = "pip install openpyxl"
        return out
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        out["status"] = "error"
        out["message"] = str(e)
        return out
    try:
        names_all = list(getattr(wb, "sheetnames", []) or [])
        names_filtered = [
            n
            for n in names_all
            if (n or "").strip().lower() not in PERF_MANUAL_SHEETS_SKIP
        ]
        names = names_filtered[:max_sheets]
        truncated_book = len(names_filtered) > max_sheets
        for sheet_name in names:
            ws = wb[sheet_name]
            raw: list[list[str]] = []
            truncated_rows = False
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= max_rows_per_sheet:
                    truncated_rows = True
                    break
                cells = [_cell_str(c) for c in row[:max_cols]]
                raw.append(cells)
            headers, body = _normalize_grid(raw, max_cols)
            if headers or body:
                out["sheets"].append(
                    {
                        "title": sheet_name,
                        "headers": headers,
                        "rows": body,
                        "truncated_rows": truncated_rows,
                    }
                )
        if truncated_book:
            out["message"] = (
                f"仅展示前 {max_sheets} 个工作表（已排除 raw 表；其余共 {len(names_filtered)} 个）。"
            )
    finally:
        wb.close()
    return out
